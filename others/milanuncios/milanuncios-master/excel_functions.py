from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment


def excel_writing(filename: str, lst: list) -> None:
    """Fills the excel file"""
    # working with file filename ('Example.xlsm')
    book = load_workbook(filename=filename, read_only=False, keep_vba=True)
    sheet = book.active
    # next_row - a string to be added to the end of an existing table of the file filename
    next_row = sheet.max_row + 1
    # variable for formatting the added string to the current sheet of the file filename
    medium_border = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='medium'),
                           bottom=Side(style='medium'))
    # if the date of the first cell of the last row of the active sheet of the file
    # filename is different from the current date
    if str(sheet[sheet.max_row][0].value)[0:10] != str(date.today()):
        # add the current date to the beginning of the list final_values_list
        lst.insert(0, date.today())
        # add a new line to the active sheet of the filename file after the last one
        sheet.append(lst)
        # set the formatting of the added line in the active sheet of the filename file
        sheet.row_dimensions[next_row].height = 23.6
        sheet[next_row][0].fill = PatternFill(fill_type='solid', start_color='FFB6FCC5')
        sheet[next_row][9].fill = PatternFill(fill_type='solid', start_color='FFF2F2F2')
        sheet[next_row][10].fill = PatternFill(fill_type='solid', start_color='FFF2F2F2')
        sheet[next_row][0].font = Font(size=10)
        for i in range(1, 12):
            sheet.cell(row=next_row, column=i).border = medium_border
            sheet.cell(row=next_row, column=i).alignment = Alignment(
                horizontal='center', vertical='center')
        for i in range(1, 11):
            sheet[next_row][i].font = Font(size=18)
        # set the numeric date format (column A) as in the previous cell on top
        sheet[next_row][0].number_format = sheet[next_row - 1][0].number_format
    # save the changes to the file
    book.save(filename)